"""
==============================================================================
Автоматизация еженедельного отчета по лидам: Bitrix24 -> Google Sheets
==============================================================================

Бизнес-логика:
- Отчетный период: Среда (00:00) — Вторник (23:59)
- Правило А: Все лиды, созданные Ср-Вт, разбитые по дням
- Правило Б: Целевые лиды Ср-Пн + «хвосты» прошлого Вторника (Вт прошлой недели
  если они стали целевыми в текущем периоде), без Вт текущего периода

Автор: сгенерировано Antigravity AI
==============================================================================
"""

import os
import json
import logging
import requests
from datetime import datetime, timedelta, date, timezone
from zoneinfo import ZoneInfo  # Python 3.9+
from google.oauth2 import service_account
from googleapiclient.discovery import build

# ---------------------------------------------------------------------------
# Настройка логирования
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Константы — ЗАМЕНИТЕ ПОД СВОЙ БИТРИКС24
# ---------------------------------------------------------------------------

# ── Маппинг SOURCE_ID -> Категория ─────────────────────────────────────────
# Сгенерировано автоматически на основе вашего списка источников из CRM
SOURCE_MAP: dict[str, str] = {
    # Я.Директ
    "ADVERTISING":      "Я.Директ",

    # SEO
    "WEB":              "SEO",          # Веб-сайт
    "10":               "SEO",          # Сайт nat-advance.ru
    "11":               "SEO",          # Сайт barssport.com
    "UC_OKOQMS":        "SEO",          # Сайт na-corporate.ru

    # Вход. звонок
    "CALL":             "Вход. звонок", # Звонок
    "CALLBACK":         "Вход. звонок", # Обратный звонок
    "8":                "Вход. звонок", # Звонок на номер: 78005007990
    "12":               "Вход. звонок", # Звонок на номер: 78432126367
    "13":               "Вход. звонок", # Звонок на номер: 78432126361
    "14":               "Вход. звонок", # Звонок на номер: 79952225983
    "15":               "Вход. звонок", # Звонок на номер: 78432126364
    "16":               "Вход. звонок", # Звонок на номер: 74991134939
    "17":               "Вход. звонок", # Звонок на номер: 74993467174
    "18":               "Вход. звонок", # Звонок на номер: 74993502734
    "19":               "Вход. звонок", # Звонок на номер: 74994041625
    "20":               "Вход. звонок", # Звонок на номер: 74994041943
    "21":               "Вход. звонок", # Звонок на номер: 74994907608
    "22":               "Вход. звонок", # Звонок на номер: 74991133791
    "23":               "Вход. звонок", # Звонок на номер: 79315210658

    # Авито
    "4|AVITO":          "Авито",        # Avito - Avito
}

# Строки таблицы (фиксированный порядок — как в сквозной аналитике Битрикс24)
TABLE_ROWS: list[str] = [
    "ВКонтакте",
    "Я.Директ e-17479930",
    "Прочий трафик",
    "Email Маркетинг",
    "Карты",
    "Avito",
    "SEO barssport.com",
    "SEO nat-advance.ru",
    "No Cookie",
    "Входящий звонок"
]

# Статусы мусорных лидов, которые не должны учитываться при подсчете общего количества лидов (Правило А)
EXCLUDED_STATUSES: set[str] = {
    "3",  # Спам/Рассылка
    "4",  # Дубль
    "5",  # Тест
    "6",  # Клиент существует в базе
    "7",  # Неверный номер
}

# Статусы лидов, которые НЕ считаются целевыми (любые другие статусы считаются целевыми)
NON_TARGET_STATUSES: set[str] = {
    "3",     # Спам/Рассылка
    "4",     # Дубль
    "5",     # Тест
    "6",     # Клиент существует в базе
    "7",     # Неверный номер
    "8",     # Ребенок
    "17",    # Не берет трубку
    "18",    # Не наш ассортимент
    "JUNK",  # Заявку не оставлял
}

# ── Временная зона ───────────────────────────────────────────────────────────
# Измените на вашу, если CRM работает в другой зоне
TZ = ZoneInfo("Europe/Moscow")

# ---------------------------------------------------------------------------
# Чтение переменных среды (GitHub Secrets)
# ---------------------------------------------------------------------------

def get_env(key: str) -> str:
    """Получает обязательную переменную среды или завершает программу с ошибкой."""
    value = os.environ.get(key)
    if not value:
        raise EnvironmentError(
            f"Переменная среды '{key}' не задана. "
            "Убедитесь, что она добавлена в GitHub Secrets."
        )
    return value


# ---------------------------------------------------------------------------
# Вычисление временных окон
# ---------------------------------------------------------------------------

def compute_date_windows(today: date | None = None) -> dict:
    """
    Вычисляет все временные окна для отчёта.

    Аргументы:
        today: Дата запуска (среда). Если None — берётся сегодня.

    Возвращает словарь с ключами:
        current_wed       — Среда текущего отчётного периода
        current_tue       — Вторник, которым закрывается период
        prev_tue_start    — Начало прошлого Вторника (дата 2 недели назад)
        prev_tue_end      — Конец прошлого Вторника
        days_of_week      — Упорядоченный список (date, label) по дням недели
    """
    if today is None:
        today = datetime.now(tz=TZ).date()

    # Если сегодня не среда — находим ближайшую прошедшую среду
    # weekday(): Пн=0, Вт=1, Ср=2, Чт=3, Пт=4, Сб=5, Вс=6
    days_since_wed = (today.weekday() - 2) % 7
    current_wed = today - timedelta(days=days_since_wed)

    # Вторник = Среда + 6 дней назад + 7 дней = Среда - 1 день
    current_tue = current_wed + timedelta(days=6)

    # Вторник прошлой недели = 7 дней до current_tue
    prev_tue = current_tue - timedelta(days=7)

    log.info("Отчётный период: %s (Ср) — %s (Вт)", current_wed, current_tue)
    log.info("Прошлый вторник (хвосты): %s", prev_tue)

    # Формируем список дней для колонок таблицы
    # Порядок: Ср, Чт, Пт, Сб+Вс (объединены), Пн, Вт
    days = []
    for offset, label in [
        (0, "Ср"),   # Среда
        (1, "Чт"),   # Четверг
        (2, "Пт"),   # Пятница
        (3, "Сб"),   # Суббота  ─┐ если Сб/Вс объединены в одну колонку,
        (4, "Вс"),   # Воскресенье ┘ скрипт суммирует их
        (5, "Пн"),   # Понедельник
        (6, "Вт"),   # Вторник
    ]:
        days.append((current_wed + timedelta(days=offset), label))

    return {
        "current_wed": current_wed,
        "current_tue": current_tue,
        "prev_tue": prev_tue,
        "days": days,  # [(date, label), ...]
    }


def to_bitrix_dt(d: date, end_of_day: bool = False) -> str:
    """
    Конвертирует дату Python в строку формата, ожидаемого Битрикс24 API.
    Битрикс24 принимает ISO 8601: "2026-05-27T00:00:00+03:00"
    """
    t = datetime(d.year, d.month, d.day, tzinfo=TZ)
    if end_of_day:
        t = t.replace(hour=23, minute=59, second=59)
    return t.isoformat()


# ---------------------------------------------------------------------------
# Запросы к Битрикс24 API
# ---------------------------------------------------------------------------

def bitrix_request(webhook_url: str, method: str, params: dict) -> list[dict]:
    """
    Выполняет запрос к REST API Битрикс24 с автоматической пагинацией.

    Битрикс24 возвращает максимум 50 записей за раз.
    Функция итерирует по страницам и возвращает полный список.
    """
    results = []
    start = 0

    while True:
        payload = {**params, "start": start}
        url = f"{webhook_url.rstrip('/')}/{method}"

        try:
            resp = requests.post(url, json=payload, timeout=30)
            resp.raise_for_status()
        except requests.RequestException as e:
            log.error("Ошибка запроса к Битрикс24: %s", e)
            raise

        data = resp.json()

        if "error" in data:
            raise RuntimeError(
                f"Ошибка API Битрикс24: {data.get('error_description', data['error'])}"
            )

        batch = data.get("result", [])
        results.extend(batch)

        total = int(data.get("total", 0))
        start += len(batch)

        log.debug("Получено %d / %d лидов", len(results), total)

        if start >= total or not batch:
            break

    return results


def fetch_leads(
    webhook_url: str,
    date_from: date,
    date_to: date,
    extra_filter: dict | None = None,
) -> list[dict]:
    """
    Запрашивает список лидов из Битрикс24 за указанный период.

    Аргументы:
        webhook_url  — URL входящего вебхука
        date_from    — Начало периода (включительно)
        date_to      — Конец периода (включительно)
        extra_filter — Дополнительные фильтры (например, по статусу)

    Возвращает список лидов (каждый — словарь с полями API).
    """
    # Поля, которые запрашиваем у API
    # SOURCE_ID — источник лида (для маппинга)
    # STATUS_ID — статус лида (для фильтрации целевых)
    # DATE_CREATE — дата создания
    # DATE_MODIFY — дата последнего изменения (нужна для хитрой логики Вт)
    # UTM_SOURCE — UTM-метка источника (запасной вариант)
    # UF_CRM_TRACKING_SOURCE_TXT — Источник сквозной аналитики (приоритетный вариант)
    fields = ["ID", "SOURCE_ID", "STATUS_ID", "DATE_CREATE", "DATE_MODIFY", "UTM_SOURCE", "UF_CRM_TRACKING_SOURCE_TXT"]

    filter_params: dict = {
        ">=DATE_CREATE": to_bitrix_dt(date_from),
        "<=DATE_CREATE": to_bitrix_dt(date_to, end_of_day=True),
    }

    if extra_filter:
        filter_params.update(extra_filter)

    params = {
        "filter": filter_params,
        "select": fields,
        "order": {"DATE_CREATE": "ASC"},
    }

    log.info(
        "Запрос лидов: %s — %s | доп. фильтр: %s",
        date_from, date_to, extra_filter or "нет",
    )

    leads = bitrix_request(webhook_url, "crm.lead.list", params)
    log.info("Получено %d лидов", len(leads))
    return leads


# ---------------------------------------------------------------------------
# Бизнес-логика: агрегация лидов
# ---------------------------------------------------------------------------

def get_source_category(lead: dict) -> str:
    """
    Определяет категорию строки таблицы по Источнику сквозной аналитики (UF_CRM_TRACKING_SOURCE_TXT).
    """
    tracking_src = str(lead.get("UF_CRM_TRACKING_SOURCE_TXT") or "").strip().lower()

    if tracking_src:
        if "я.директ" in tracking_src or "direct" in tracking_src or "yandex" in tracking_src:
            return "Я.Директ e-17479930"
        if "seo barssport.com" in tracking_src or "barssport.com" in tracking_src:
            return "SEO barssport.com"
        if "seo nat-advance.ru" in tracking_src or "nat-advance.ru" in tracking_src:
            return "SEO nat-advance.ru"
        if "google ads" in tracking_src or "google_ads" in tracking_src:
            return "Прочий трафик"  # Google Ads удален из таблицы, направляем в Прочий трафик
        if "вконтакте" in tracking_src or "vk" in tracking_src:
            return "ВКонтакте"
        if "звонок" in tracking_src or "call" in tracking_src:
            return "Входящий звонок"
        if "avito" in tracking_src or "авито" in tracking_src:
            return "Avito"
        if "email" in tracking_src or "рассылка" in tracking_src:
            return "Email Маркетинг"
        if "карты" in tracking_src or "maps" in tracking_src:
            return "Карты"
        if "no cookie" in tracking_src:
            return "No Cookie"
        if "прочий" in tracking_src or "other" in tracking_src:
            return "Прочий трафик"

        # Если не подошло под ключевые слова, но совпадает с одной из строк таблицы
        for row in TABLE_ROWS:
            if row.lower() == tracking_src:
                return row

    # Резервный вариант по SOURCE_ID (для старых или ручных лидов)
    source_id = lead.get("SOURCE_ID", "") or ""
    utm_source = lead.get("UTM_SOURCE", "") or ""

    if source_id in SOURCE_MAP:
        mapped = SOURCE_MAP[source_id]
        # Сопоставляем старые названия с новыми
        mapping_old_to_new = {
            "Я.Директ": "Я.Директ e-17479930",
            "SEO": "SEO barssport.com",
            "Вход. звонок": "Входящий звонок",
            "Авито": "Avito"
        }
        return mapping_old_to_new.get(mapped, "Прочий трафик")

    # Если SOURCE_ID не в маппинге — проверяем UTM_SOURCE
    utm_lower = utm_source.lower()
    if "yandex" in utm_lower or "direct" in utm_lower or "ya" in utm_lower:
        return "Я.Директ e-17479930"
    if "avito" in utm_lower:
        return "Avito"
    if "seo" in utm_lower or "organic" in utm_lower or "google" in utm_lower:
        return "SEO barssport.com"

    return "Прочий трафик"


def aggregate_leads(windows: dict, all_leads: list[dict]) -> dict:
    """
    Агрегирует лиды по дням недели и целевому статусу.

    Считает все созданные лиды за период Ср-Вт по дням и источникам (без EXCLUDED_STATUSES).
    Целевые лиды считаются из этих же лидов (статус не в NON_TARGET_STATUSES).

    Возвращает структуру:
    {
        "daily": {
            "Я.Директ": {"Ср": 6, "Чт": 5, "Пт": 5, "Сб": 0, "Вс": 0, "Пн": 18, "Вт": 7, "Итого": 41},
            ...
        },
        "targeted": {"Я.Директ": 25, "SEO": 15, ...},
    }
    """
    days = windows["days"]          # [(date, label), ...]
    current_wed = windows["current_wed"]
    current_tue = windows["current_tue"]

    # Словарь date -> label для быстрого поиска
    date_to_label: dict[date, str] = {d: lbl for d, lbl in days}

    # Инициализируем структуру результата
    daily: dict[str, dict[str, int]] = {}
    targeted: dict[str, int] = {}

    for row in TABLE_ROWS:
        daily[row] = {lbl: 0 for _, lbl in days}
        daily[row]["Итого"] = 0
        targeted[row] = 0

    for lead in all_leads:
        status = lead.get("STATUS_ID", "")
        raw_date = lead.get("DATE_CREATE", "")

        # Дата создания из API: "2026-05-27T14:30:00+03:00"
        try:
            created_dt = datetime.fromisoformat(raw_date).astimezone(TZ)
            created_date = created_dt.date()
        except (ValueError, TypeError):
            log.warning("Неверный формат DATE_CREATE у лида %s: %s", lead.get("ID"), raw_date)
            continue

        category = get_source_category(lead)
        if category not in daily:
            continue  # Неизвестный источник — пропускаем

        # Проверяем, попадает ли дата создания в наш отчетный период (Ср-Вт)
        label = date_to_label.get(created_date)
        if label:
            # 1. Считаем лид в общее количество (если он не мусорный)
            if status not in EXCLUDED_STATUSES:
                daily[category][label] += 1
                daily[category]["Итого"] += 1

            # 2. Считаем лид в целевые (если статус целевой)
            if status not in NON_TARGET_STATUSES:
                targeted[category] += 1

    log.info("Агрегация завершена. Целевые: %s", targeted)
    return {"daily": daily, "targeted": targeted}


# ---------------------------------------------------------------------------
# Google Sheets: вспомогательные функции, поиск нужного блока и запись данных
# ---------------------------------------------------------------------------

def get_sheets_service(service_account_json: str):
    """Создаёт авторизованный клиент Google Sheets API."""
    creds_dict = json.loads(service_account_json)
    creds = service_account.Credentials.from_service_account_info(
        creds_dict,
        scopes=["https://www.googleapis.com/auth/spreadsheets"],
    )
    service = build("sheets", "v4", credentials=creds)
    return service.spreadsheets()


def get_or_create_sheet_tab(sheets, spreadsheet_id: str, tab_name: str):
    """
    Проверяет, существует ли вкладка с именем tab_name.
    Если нет — создаёт её.
    Возвращает sheet_id вкладки.
    """
    meta = sheets.get(spreadsheetId=spreadsheet_id).execute()
    existing_tabs = [sheet["properties"]["title"] for sheet in meta.get("sheets", [])]
    log.info("Доступные вкладки в таблице: %s", existing_tabs)
    for sheet in meta.get("sheets", []):
        if sheet["properties"]["title"] == tab_name:
            log.info("Вкладка '%s' найдена.", tab_name)
            return sheet["properties"]["sheetId"]

    # Вкладка не найдена — создаём
    log.info("Вкладка '%s' не найдена. Создаём...", tab_name)
    body = {
        "requests": [{
            "addSheet": {
                "properties": {"title": tab_name}
            }
        }]
    }
    resp = sheets.batchUpdate(spreadsheetId=spreadsheet_id, body=body).execute()
    return resp["replies"][0]["addSheet"]["properties"]["sheetId"]


def find_week_block_row(sheets, spreadsheet_id: str, tab_name: str, target_wed: date) -> int | None:
    """
    Ищет начальную строку блока для текущей недели в таблице Google Sheets.
    """
    allowed_date_strs = [
        target_wed.strftime("%d.%m"),                           # "06.05"
        f"{target_wed.day}.{target_wed.month:02d}",             # "6.05" (без ведущего нуля)
        target_wed.strftime("%d.%m.%Y"),                        # "06.05.2026"
        f"{target_wed.day}.{target_wed.month:02d}.{target_wed.year}" # "6.05.2026"
    ]

    range_notation = f"'{tab_name}'!A1:M200"
    try:
        result = sheets.values().get(
            spreadsheetId=spreadsheet_id,
            range=range_notation,
        ).execute()
        rows = result.get("values", [])
        for row_idx, row in enumerate(rows):
            for cell in row:
                cell_str = str(cell).strip()
                if cell_str in allowed_date_strs:
                    log.info(
                        "Дата недели найдена в строке %d (значение: '%s')",
                        row_idx + 1, cell_str,
                    )
                    return row_idx + 1  # 1-indexed
    except Exception as e:
        log.warning("Ошибка при чтении листа '%s': %s", tab_name, e)

    log.warning(
        "Блок для недели %s не найден на листе '%s'. Проверяем все остальные листы...",
        target_wed, tab_name,
    )

    try:
        meta = sheets.get(spreadsheetId=spreadsheet_id).execute()
        for sheet in meta.get("sheets", []):
            title = sheet["properties"]["title"]
            if title == tab_name:
                continue
            res = sheets.values().get(
                spreadsheetId=spreadsheet_id,
                range=f"'{title}'!A1:M200",
            ).execute()
            r_rows = res.get("values", [])
            for r_idx, r in enumerate(r_rows):
                for cell in r:
                    cell_str = str(cell).strip()
                    if cell_str in allowed_date_strs:
                        log.info(
                            "💡 НАЙДЕНО на листе '%s' в строке %d (значение: '%s')",
                            title, r_idx + 1, cell_str,
                        )
    except Exception as e:
        log.warning("Не удалось выполнить поиск по всем листам: %s", e)

    return None


def hex_to_rgb(hex_str: str) -> dict:
    """Конвертирует HEX цвет в RGB формат для Google Sheets API (значения от 0.0 до 1.0)."""
    hex_str = hex_str.lstrip('#')
    if len(hex_str) == 3:
        hex_str = ''.join([c*2 for c in hex_str])
    r = int(hex_str[0:2], 16) / 255.0
    g = int(hex_str[2:4], 16) / 255.0
    b = int(hex_str[4:6], 16) / 255.0
    return {"red": r, "green": g, "blue": b}


def make_border(hex_color: str = "#CBD5E1") -> dict:
    """Создаёт объект границы для Google Sheets API."""
    return {
        "style": "SOLID",
        "color": hex_to_rgb(hex_color)
    }


def create_month_template(sheets, spreadsheet_id: str, tab_name: str, sheet_id: int, current_wed: date):
    """
    Создаёт шаблон со всеми неделями месяца на указанной вкладке с профессиональным дизайном и легендой.
    """
    log.info("Шаблон не найден. Создаём структуру недель для листа '%s'...", tab_name)

    # 1. Находим все Среды в текущем месяце
    year = current_wed.year
    month = current_wed.month

    wednesdays = []
    d = date(year, month, 1)
    while d.month == month:
        if d.weekday() == 2:  # Wednesday
            wednesdays.append(d)
        d += timedelta(days=1)

    log.info("Найдено %d недель в месяце: %s", len(wednesdays), [w.strftime("%d.%m") for w in wednesdays])

    # Справка-легенда занимает первые 5 строк (0..4).
    # Индекс 5 - пустая строка-разделитель.
    # Каждая неделя занимает 28 строк (с отступами).
    week_height = 2 + len(TABLE_ROWS) + 1 + 1 + 1 + len(TABLE_ROWS) + 1 + 1 + 1
    num_rows = 6 + len(wednesdays) * week_height
    values = [["" for _ in range(14)] for _ in range(num_rows)]
    
    # Список запросов batchUpdate (для слияний и форматирования)
    requests_list = []

    # ── Справка по логике отчёта (Легенда) ─────────────────────────────────
    values[0][0] = "Справка по логике отчёта (Сквозная аналитика)"
    values[1][0] = "• Период отчёта: Среда (00:00) — Вторник (23:59). Лиды за выходные (Сб и Вс) автоматически суммируются в колонку Понедельника."
    values[2][0] = "• Исключено из общего количества лидов (Дубли, Спам, Тесты, Неверные номера, Клиент существует в базе)."
    values[3][0] = "• Исключено из целевых лидов (Все вышеперечисленное + Ребенок, Не берет трубку, Не наш ассортимент, Заявку не оставлял)."
    values[4][0] = "• Расчет стоимости лида: Стоимость рассчитывается автоматически по формуле сразу после ручного ввода расходов в колонку 'Израсходованный бюджет'."

    def add_merge(s_r, e_r, s_c, e_c):
        requests_list.append({
            "mergeCells": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": s_r,
                    "endRowIndex": e_r + 1,
                    "startColumnIndex": s_c,
                    "endColumnIndex": e_c + 1,
                },
                "mergeType": "MERGE_ALL",
            }
        })

    def format_range(s_r, e_r, s_c, e_c, bg_hex=None, fg_hex=None, size=9, bold=False, italic=False, align="CENTER", num_pattern=None):
        cell_format = {}
        fields = []

        if bg_hex:
            cell_format["backgroundColor"] = hex_to_rgb(bg_hex)
            fields.append("backgroundColor")

        text_fmt = {}
        if fg_hex:
            text_fmt["foregroundColor"] = hex_to_rgb(fg_hex)
        text_fmt["fontSize"] = size
        text_fmt["bold"] = bold
        text_fmt["italic"] = italic
        cell_format["textFormat"] = text_fmt
        fields.append("textFormat")

        cell_format["horizontalAlignment"] = align
        cell_format["verticalAlignment"] = "MIDDLE"
        fields.extend(["horizontalAlignment", "verticalAlignment"])

        if num_pattern:
            cell_format["numberFormat"] = {
                "type": "CURRENCY" if "₽" in num_pattern else ("PERCENT" if "%" in num_pattern else "NUMBER"),
                "pattern": num_pattern
            }
            fields.append("numberFormat")

        border = make_border("#CBD5E1")
        cell_format["borders"] = {
            "top": border,
            "bottom": border,
            "left": border,
            "right": border
        }
        fields.append("borders")

        requests_list.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": s_r,
                    "endRowIndex": e_r + 1,
                    "startColumnIndex": s_c,
                    "endColumnIndex": e_c + 1,
                },
                "cell": {
                    "userEnteredFormat": cell_format
                },
                "fields": f"userEnteredFormat({','.join(fields)})"
            }
        })

    # Слияния и стили для Легенды
    add_merge(0, 0, 0, 12)
    add_merge(1, 1, 0, 12)
    add_merge(2, 2, 0, 12)
    add_merge(3, 3, 0, 12)
    add_merge(4, 4, 0, 12)
    
    format_range(0, 0, 0, 12, bg_hex="#1E293B", fg_hex="#FFFFFF", size=11, bold=True)
    format_range(1, 4, 0, 12, bg_hex="#F8FAFC", size=9, align="LEFT")
    format_range(4, 4, 0, 12, bg_hex="#F8FAFC", size=9, align="LEFT", italic=True)

    for idx, wed in enumerate(wednesdays):
        start_row = 6 + idx * week_height

        # Вычисляем даты
        thu = wed + timedelta(days=1)
        fri = wed + timedelta(days=2)
        mon = wed + timedelta(days=5)
        tue = wed + timedelta(days=6)

        d_wed = wed.strftime("%d.%m")
        d_thu = thu.strftime("%d.%m")
        d_fri = fri.strftime("%d.%m")
        d_mon = mon.strftime("%d.%m")
        d_tue = tue.strftime("%d.%m")

        # ── Первая таблица: Данные по лидам ─────────────────────────────────
        # Заголовки (строка start_row)
        values[start_row][0] = "Источник"
        values[start_row][1] = "Период"
        values[start_row][6] = "Итого за неделю"
        values[start_row][7] = "Кол-во целевых"
        values[start_row][8] = "Израсходованный бюджет"
        values[start_row][9] = "Общая цена лида"
        values[start_row][10] = "Цена целевого лида"
        values[start_row][11] = "Динамика количества целевого лида относительной прошлой недели, %"
        values[start_row][12] = "Динамика стоимости целевого лида относительной прошлой недели, %"

        # Подзаголовки дат (строка start_row + 1)
        values[start_row+1][1] = d_wed
        values[start_row+1][2] = d_thu
        values[start_row+1][3] = d_fri
        values[start_row+1][4] = d_mon
        values[start_row+1][5] = d_tue

        # Источники
        for s_idx, src in enumerate(TABLE_ROWS):
            r_idx = start_row + 2 + s_idx
            values[r_idx][0] = src

            # Формулы цены лида: Общая = Бюджет/Лиды, Целевая = Бюджет/Целевые
            r_num = r_idx + 1  # 1-indexed row number in sheet
            values[r_idx][9] = f'=IF(G{r_num}>0; I{r_num}/G{r_num}; "")'
            values[r_idx][10] = f'=IF(H{r_num}>0; I{r_num}/H{r_num}; "")'

        # Строка ИТОГО
        itog_idx = start_row + 2 + len(TABLE_ROWS)
        r_total = itog_idx + 1
        values[itog_idx][0] = "ИТОГО"
        for col_idx in range(1, 9):
            col_letter = chr(65 + col_idx)  # 65 = 'A'
            start_r = r_total - len(TABLE_ROWS)
            end_r = r_total - 1
            values[itog_idx][col_idx] = f'=SUM({col_letter}{start_r}:{col_letter}{end_r})'

        # ── Вторая таблица: Комментарии и планы ─────────────────────────────
        comments_header_idx = itog_idx + 2
        values[comments_header_idx][0] = "Источник"
        values[comments_header_idx][1] = "Комментарий по результату недели"
        values[comments_header_idx][2] = "План мероприятий на следующую неделю"

        for s_idx, src in enumerate(TABLE_ROWS):
            r_c = comments_header_idx + 1 + s_idx
            values[r_c][0] = src

        proposal_idx = comments_header_idx + 1 + len(TABLE_ROWS) + 1
        values[proposal_idx][0] = "Предложение на тест новой площадки:"

        # ── Слияния заголовков ──
        add_merge(start_row, start_row + 1, 0, 0)   # Источник
        add_merge(start_row, start_row + 1, 6, 6)   # Итого за неделю
        add_merge(start_row, start_row + 1, 7, 7)   # Кол-во целевых
        add_merge(start_row, start_row + 1, 8, 8)   # Бюджет
        add_merge(start_row, start_row + 1, 9, 9)   # Общая цена лида
        add_merge(start_row, start_row + 1, 10, 10) # Цена целевого лида
        add_merge(start_row, start_row + 1, 11, 11) # Динамика %
        add_merge(start_row, start_row + 1, 12, 12) # Динамика %
        add_merge(start_row, start_row, 1, 5)        # Period

        # ── Стили заголовков первой таблицы ──
        format_range(start_row, start_row, 0, 12, bg_hex="#1E293B", fg_hex="#FFFFFF", size=10, bold=True)
        format_range(start_row + 1, start_row + 1, 0, 12, bg_hex="#334155", fg_hex="#FFFFFF", size=9, bold=True)

        # ── Стили строк данных ──
        for s_idx, src in enumerate(TABLE_ROWS):
            r_idx = start_row + 2 + s_idx
            bg = "#FFFFFF" if s_idx % 2 == 0 else "#F8FAFC"

            format_range(r_idx, r_idx, 0, 12, bg_hex=bg, size=9)
            format_range(r_idx, r_idx, 0, 0, bg_hex=bg, size=9, bold=True, align="LEFT")
            format_range(r_idx, r_idx, 1, 5, bg_hex=bg, size=9, num_pattern="#,##0")
            format_range(r_idx, r_idx, 6, 6, bg_hex=bg, size=9, bold=True, num_pattern="#,##0")
            format_range(r_idx, r_idx, 7, 7, bg_hex="#ECFDF5", size=9, bold=True, num_pattern="#,##0")
            format_range(r_idx, r_idx, 8, 8, bg_hex="#FEF3C7", size=9, align="RIGHT", num_pattern="#,##0\" ₽\"")
            format_range(r_idx, r_idx, 9, 10, bg_hex=bg, size=9, align="RIGHT", num_pattern="#,##0\" ₽\"")
            format_range(r_idx, r_idx, 11, 12, bg_hex=bg, size=9, num_pattern="0.0%")

        # ── Стили строки ИТОГО ──
        format_range(itog_idx, itog_idx, 0, 12, bg_hex="#E2E8F0", size=9, bold=True)
        format_range(itog_idx, itog_idx, 0, 0, bg_hex="#E2E8F0", size=9, bold=True, align="LEFT")
        format_range(itog_idx, itog_idx, 1, 7, bg_hex="#E2E8F0", size=9, bold=True, num_pattern="#,##0")
        format_range(itog_idx, itog_idx, 8, 8, bg_hex="#E2E8F0", size=9, bold=True, align="RIGHT", num_pattern="#,##0\" ₽\"")

        # ── Стили второй таблицы (Комментарии и планы) ──
        add_merge(comments_header_idx, comments_header_idx, 1, 6)
        add_merge(comments_header_idx, comments_header_idx, 7, 12)
        for s_idx in range(len(TABLE_ROWS)):
            r_c = comments_header_idx + 1 + s_idx
            add_merge(r_c, r_c, 1, 6)
            add_merge(r_c, r_c, 7, 12)
        add_merge(proposal_idx, proposal_idx, 0, 12)

        format_range(comments_header_idx, comments_header_idx, 0, 12, bg_hex="#475569", fg_hex="#FFFFFF", size=9, bold=True)
        format_range(comments_header_idx, comments_header_idx, 0, 0, bg_hex="#475569", fg_hex="#FFFFFF", size=9, bold=True, align="LEFT")

        for s_idx, src in enumerate(TABLE_ROWS):
            r_c = comments_header_idx + 1 + s_idx
            bg = "#FFFFFF" if s_idx % 2 == 0 else "#F8FAFC"
            format_range(r_c, r_c, 0, 12, bg_hex=bg, size=9)
            format_range(r_c, r_c, 0, 0, bg_hex=bg, size=9, bold=True, align="LEFT")

        format_range(proposal_idx, proposal_idx, 0, 12, bg_hex="#F8FAFC", size=9, align="LEFT", italic=True)

    # ── 3. Установка высоты строк ─────────────────────────────────────────
    for r in range(6):
        requests_list.append({
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": r,
                    "endIndex": r + 1
                },
                "properties": {"pixelSize": 24},
                "fields": "pixelSize"
            }
        })

    for idx in range(len(wednesdays)):
        start_row = 6 + idx * week_height
        requests_list.append({
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": start_row,
                    "endIndex": start_row + 1
                },
                "properties": {"pixelSize": 35},
                "fields": "pixelSize"
            }
        })
        requests_list.append({
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": start_row + 1,
                    "endIndex": start_row + 2
                },
                "properties": {"pixelSize": 24},
                "fields": "pixelSize"
            }
        })
        for r_offset in range(len(TABLE_ROWS) + 1):  # Data + ИТОГО
            r_idx = start_row + 2 + r_offset
            requests_list.append({
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "ROWS",
                        "startIndex": r_idx,
                        "endIndex": r_idx + 1
                    },
                    "properties": {"pixelSize": 24},
                    "fields": "pixelSize"
                }
            })
        c_header = start_row + 2 + len(TABLE_ROWS) + 2
        requests_list.append({
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": c_header,
                    "endIndex": c_header + 1
                },
                "properties": {"pixelSize": 28},
                "fields": "pixelSize"
            }
        })
        for r_offset in range(len(TABLE_ROWS) + 1):  # Comments + Proposal
            r_idx = c_header + 1 + r_offset
            requests_list.append({
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "ROWS",
                        "startIndex": r_idx,
                        "endIndex": r_idx + 1
                    },
                    "properties": {"pixelSize": 24},
                    "fields": "pixelSize"
                }
            })

    # ── 4. Установка ширины столбцов ──────────────────────────────────────
    col_widths = {
        0: 180,  # Источник
        1: 70, 2: 70, 3: 70, 4: 70, 5: 70,  # Ср, Чт, Пт, Пн, Вт
        6: 120,  # Итого за неделю
        7: 120,  # Кол-во целевых
        8: 160,  # Израсходованный бюджет
        9: 140,  # Общая цена лида
        10: 140, # Цена целевого лида
        11: 140, # Динамика целевых %
        12: 140, # Динамика стоимости %
    }

    for col_idx, width in col_widths.items():
        requests_list.append({
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": col_idx,
                    "endIndex": col_idx + 1
                },
                "properties": {
                    "pixelSize": width
                },
                "fields": "pixelSize"
            }
        })

    # 5. Записываем значения на лист
    range_name = f"'{tab_name}'!A1"
    body = {"values": values}
    sheets.values().update(
        spreadsheetId=spreadsheet_id,
        range=range_name,
        valueInputOption="USER_ENTERED",
        body=body,
    ).execute()

    # 6. Отправляем слияния и форматирование
    if requests_list:
        sheets.batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": requests_list},
        ).execute()

    log.info("Шаблон на лист '%s' успешно записан.", tab_name)


def parse_direct_csv_files(date_from: date, date_to: date) -> float | None:
    """
    Ищет CSV-файлы статистики Яндекс.Директ в корне репозитория.

    Формат имени файла: *_e-17479930.csv  (как выгружает Яндекс.Директ)
    Формат содержимого: CSV с заголовком, колонка «Расход, ₽», даты в «День» dd.mm.yyyy.

    Суммирует расходы за строки, где дата попадает в [date_from, date_to].
    Если подходящих файлов нет — возвращает None.
    Если данных за нужный период нет — возвращает 0.0.
    Может обрабатывать несколько файлов (складывает).
    """
    import glob
    import csv as _csv

    # Ищем все CSV-файлы с данными Директа в корне репозитория
    repo_root = os.path.dirname(os.path.abspath(__file__))
    patterns = [
        os.path.join(repo_root, "*_e-17479930.csv"),
        os.path.join(repo_root, "direct_*.csv"),
        os.path.join(repo_root, "yandex_direct_*.csv"),
    ]
    found_files = []
    for pattern in patterns:
        found_files.extend(glob.glob(pattern))
    found_files = list(set(found_files))  # убираем дубли

    if not found_files:
        log.debug("CSV-файлы Яндекс.Директ не найдены в корне репозитория.")
        return None

    log.info("Найдено CSV-файлов Яндекс.Директ: %d → %s", len(found_files),
             [os.path.basename(f) for f in found_files])

    total_cost = 0.0
    found_any_row = False

    for filepath in found_files:
        try:
            with open(filepath, encoding="utf-8-sig", newline="") as fh:
                reader = _csv.DictReader(fh)

                # Находим нужные колонки (Яндекс иногда меняет регистр/пробелы)
                fieldnames = reader.fieldnames or []
                date_col  = next((c for c in fieldnames if "день" in c.lower()), None)
                cost_col  = next(
                    (c for c in fieldnames
                     if "расход" in c.lower() or "стоимость" in c.lower() or "cost" in c.lower()),
                    None,
                )

                if not date_col or not cost_col:
                    log.warning("[CSV] Не найдены нужные колонки в '%s'. Колонки: %s",
                                os.path.basename(filepath), fieldnames)
                    continue

                log.info("[CSV] Обрабатываем '%s' | дата: '%s' | расход: '%s'",
                         os.path.basename(filepath), date_col, cost_col)

                for row in reader:
                    date_val = row.get(date_col, "").strip()
                    cost_val = row.get(cost_col, "").strip()

                    # Пропускаем строку «Итого» и пустые строки
                    if not date_val or date_val.lower() in ("итого", "total", ""):
                        continue

                    # Парсим дату в формате dd.mm.yyyy
                    try:
                        row_date = date(
                            int(date_val[6:10]),
                            int(date_val[3:5]),
                            int(date_val[0:2]),
                        )
                    except (ValueError, IndexError):
                        continue

                    # Проверяем попадание в нужный период
                    if not (date_from <= row_date <= date_to):
                        continue

                    # Парсим сумму расхода
                    if not cost_val or cost_val in ("-", "--", ""):
                        continue
                    try:
                        cost_num = float(
                            cost_val.replace(",", ".").replace("\xa0", "").replace(" ", "")
                        )
                        total_cost += cost_num
                        found_any_row = True
                    except ValueError:
                        continue

        except Exception as exc:
            log.warning("[CSV] Ошибка при чтении '%s': %s", os.path.basename(filepath), exc)
            continue

    if not found_any_row:
        log.warning("[CSV] Файлы найдены, но данных за период %s—%s нет.", date_from, date_to)
        return None

    log.info("[CSV] Расходы из CSV за период %s—%s: %.2f руб.", date_from, date_to, total_cost)
    return total_cost


def fetch_yandex_direct_cost(
    token: str,
    client_login: str | None,
    date_from: date,
    date_to: date,
) -> float | None:
    """
    Запрашивает фактический расход из API Яндекс.Директа (Reports Service v5)
    за указанный период и возвращает сумму в рублях (с НДС).
    Если данных нет или произошла ошибка, возвращает None.
    """
    import time
    token = token.strip()
    if client_login:
        client_login = client_login.strip()
    log.info("Запрос расходов Яндекс.Директ с %s по %s...", date_from, date_to)
    
    url = "https://api.direct.yandex.com/v5/reports"
    
    # Заголовки авторизации и параметров
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept-Language": "ru",
        "processingMode": "auto",
        "returnMoneyInMicros": "false",  # Возвращать в рублях, а не в микрокопейках
        "skipReportHeader": "true",
        "skipReportSummary": "true",
    }
    if client_login:
        headers["Client-Login"] = client_login

    # Тело запроса отчета
    payload = {
        "params": {
            "SelectionCriteria": {
                "DateFrom": date_from.isoformat(),
                "DateTo": date_to.isoformat()
            },
            "FieldNames": ["Cost"],
            "ReportName": f"Report_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}",
            "ReportType": "ACCOUNT_PERFORMANCE_REPORT",
            "DateRangeType": "CUSTOM_DATE",
            "Format": "TSV",
            "IncludeDiscount": "NO",
            "IncludeVAT": "YES"
        }
    }

    # Отправляем запрос с обработкой ожидания очереди 201/202
    for attempt in range(5):
        try:
            resp = requests.post(url, json=payload, headers=headers, timeout=30)
            if resp.status_code == 200:
                break
            elif resp.status_code in (201, 202):
                log.info("Отчет Яндекс.Директ ставится в очередь, ожидание (попытка %d/5)...", attempt + 1)
                time.sleep(5)
            else:
                log.error("Ошибка API Яндекс.Директ (Код %d): %s", resp.status_code, resp.text)
                return None
        except requests.RequestException as e:
            log.error("Сетевая ошибка при запросе к API Яндекс.Директ: %s", e)
            return None
    else:
        log.error("Не удалось дождаться отчета Яндекс.Директ.")
        return None

    # Парсим TSV ответ
    lines = resp.text.strip().split("\n")
    if not lines or len(lines) <= 1:
        log.warning("Отчет Яндекс.Директ не вернул данных о расходах.")
        return 0.0

    # Находим индекс колонки Cost
    header = lines[0].split("\t")
    cost_idx = -1
    for i, col in enumerate(header):
        if col.strip() == "Cost":
            cost_idx = i
            break

    if cost_idx == -1:
        log.error("Колонка 'Cost' не найдена в ответе API Яндекс.Директ. Заголовки: %s", header)
        return None

    # Суммируем расходы
    total_cost = 0.0
    for line in lines[1:]:
        cols = line.split("\t")
        if len(cols) > cost_idx:
            val_str = cols[cost_idx].strip()
            if val_str and val_str != "--":
                try:
                    total_cost += float(val_str)
                except ValueError:
                    log.warning("Не удалось спарсить значение расхода как число: '%s'", val_str)

    log.info("Успешно получены расходы Яндекс.Директ: %.2f руб. (с НДС)", total_cost)
    return total_cost


# ──────────────────────────────────────────────────────────────────────────────
# Вспомогательные утилиты: имитация человеческого поведения в браузере
# ──────────────────────────────────────────────────────────────────────────────

import random as _random
import time as _timeit


def _human_delay(min_ms: float = 300, max_ms: float = 900) -> None:
    """Случайная пауза в диапазоне [min_ms, max_ms] миллисекунд."""
    _timeit.sleep(_random.uniform(min_ms, max_ms) / 1000)


def _human_type(page, selector: str, text: str) -> None:
    """Вводит текст посимвольно с случайными задержками (имитация живого набора)."""
    el = page.locator(selector).first
    el.click()
    _human_delay(200, 500)
    el.fill("")
    for char in text:
        el.type(char, delay=_random.randint(60, 180))
        if _random.random() < 0.08:          # иногда делаем паузу, «думаем»
            _human_delay(200, 600)


def _human_move_and_click(page, locator) -> None:
    """Перемещает мышь к элементу со случайным смещением и кликает."""
    box = locator.bounding_box()
    if box:
        tx = box["x"] + box["width"]  * _random.uniform(0.3, 0.7)
        ty = box["y"] + box["height"] * _random.uniform(0.3, 0.7)
        page.mouse.move(tx, ty)
        _human_delay(80, 250)
    locator.click()


def _human_scroll(page) -> None:
    """Прокручивает страницу вниз, затем немного назад (как живой человек)."""
    page.mouse.wheel(0, _random.randint(200, 600))
    _human_delay(400, 800)
    page.mouse.wheel(0, -_random.randint(50, 200))


# ──────────────────────────────────────────────────────────────────────────────


def fetch_yandex_direct_cost_browser(
    yandex_login: str,
    yandex_password: str,
    secret_answer: str | None,
    date_from: date,
    date_to: date,
) -> float | None:
    """
    Скачивает статистику расходов Яндекс.Директ через браузерную автоматизацию
    (Playwright, headless Chromium).

    Полностью имитирует поведение живого пользователя:
    - случайные задержки между действиями
    - посимвольный ввод с нерегулярными паузами
    - движения мыши со случайным смещением от центра кнопки
    - прокрутка страницы для «прочтения» контента
    - скрытие navigator.webdriver / AutomationControlled

    Возвращает суммарные расходы в рублях (с НДС) или None при ошибке.
    """
    import tempfile
    import os

    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    except ImportError:
        log.error("Playwright не установлен. Проверьте requirements.txt.")
        return None

    try:
        from openpyxl import load_workbook  # noqa: F401
    except ImportError:
        log.error("openpyxl не установлен. Проверьте requirements.txt.")
        return None

    date_from_str = date_from.strftime("%d.%m.%Y")
    date_to_str   = date_to.strftime("%d.%m.%Y")
    log.info(
        "Запуск браузера (Playwright) для скачивания статистики Яндекс.Директ (%s — %s)...",
        date_from_str, date_to_str,
    )
    tmpdir = tempfile.mkdtemp()

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--disable-blink-features=AutomationControlled",
            ],
        )
        context = browser.new_context(
            locale="ru-RU",
            timezone_id="Europe/Moscow",
            viewport={"width": 1366, "height": 768},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            accept_downloads=True,
            downloads_path=tmpdir,
        )
        # Убираем признак автоматизации
        context.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )
        page = context.new_page()
        page.set_default_timeout(60_000)

        downloaded_file = None
        try:
            # ── 1. Страница входа ─────────────────────────────────────────────
            log.info("[Browser] Открываем страницу входа Яндекс...")
            page.goto("https://passport.yandex.ru/auth", wait_until="domcontentloaded")
            _human_delay(1500, 3000)
            _human_scroll(page)

            # ── 2. Логин ──────────────────────────────────────────────────────
            log.info("[Browser] Вводим логин...")
            for login_sel in ["#passp-field-login", "input[name='login']", "input[type='text']"]:
                try:
                    if page.locator(login_sel).first.is_visible(timeout=3000):
                        _human_type(page, login_sel, yandex_login)
                        break
                except Exception:
                    continue

            _human_delay(600, 1200)
            _human_move_and_click(page, page.locator("button[type='submit']").first)
            _human_delay(2000, 3500)

            # ── 3. Пароль ─────────────────────────────────────────────────────
            log.info("[Browser] Вводим пароль...")
            try:
                pw_input = page.locator("#passp-field-passwd")
                pw_input.wait_for(state="visible", timeout=15_000)
            except Exception:
                log.error("[Browser] Поле пароля не появилось. Проверьте логин.")
                return None

            _human_delay(800, 1500)
            _human_type(page, "#passp-field-passwd", yandex_password)
            _human_delay(500, 1000)
            _human_move_and_click(page, page.locator("button[type='submit']").first)
            _human_delay(2500, 4000)

            # ── 4. Секретный вопрос (если появился) ──────────────────────────
            if page.locator("#passp-field-answer").is_visible():
                if secret_answer:
                    log.info("[Browser] Вводим ответ на секретный вопрос...")
                    _human_delay(800, 1500)
                    _human_type(page, "#passp-field-answer", secret_answer)
                    _human_delay(500, 1000)
                    _human_move_and_click(page, page.locator("button[type='submit']").first)
                    _human_delay(2000, 3500)
                else:
                    log.error("[Browser] Яндекс запросил ответ на секретный вопрос, "
                              "но YANDEX_SECRET_ANSWER не задан.")
                    return None

            # ── 5. Проверка успешного входа ───────────────────────────────────
            log.info("[Browser] Текущий URL после входа: %s", page.url)
            if "passport.yandex" in page.url:
                try:
                    page.wait_for_url(
                        lambda url: "passport.yandex" not in url, timeout=12_000
                    )
                except Exception:
                    log.error("[Browser] Не удалось войти. URL: %s. "
                              "Проверьте логин/пароль и отключите 2FA.", page.url)
                    return None

            log.info("[Browser] Успешно вошли! URL: %s", page.url)

            # ── 6. Переходим в Яндекс.Директ ─────────────────────────────────
            _human_delay(1200, 2500)
            log.info("[Browser] Переходим в Яндекс.Директ...")
            page.goto("https://direct.yandex.ru", wait_until="domcontentloaded")
            _human_delay(2000, 4000)
            _human_scroll(page)

            # ── 7. Открываем Мастер отчётов ──────────────────────────────────
            log.info("[Browser] Открываем Мастер отчётов...")
            page.goto(
                "https://direct.yandex.ru/registered/main.pl?cmd=showReportWizard",
                wait_until="domcontentloaded",
            )
            _human_delay(3000, 5000)
            _human_scroll(page)

            # ── 8. Выбираем период дат ────────────────────────────────────────
            log.info("[Browser] Устанавливаем даты %s — %s...", date_from_str, date_to_str)
            for period_sel in [
                "text=Произвольный период",
                "label:has-text('Произвольный')",
                "[class*='period'] [value='custom']",
            ]:
                try:
                    loc = page.locator(period_sel).first
                    if loc.is_visible(timeout=2000):
                        _human_move_and_click(page, loc)
                        _human_delay(500, 1000)
                        break
                except Exception:
                    continue

            # Заполняем поля дат
            for candidates, value in [
                (
                    ["input[name='date_from']", "input[id*='date_from']", "input[class*='date-from']"],
                    date_from_str,
                ),
                (
                    ["input[name='date_to']", "input[id*='date_to']", "input[class*='date-to']"],
                    date_to_str,
                ),
            ]:
                for sel in candidates:
                    try:
                        loc = page.locator(sel).first
                        if loc.is_visible(timeout=2000):
                            _human_delay(300, 700)
                            loc.triple_click()
                            _human_delay(150, 300)
                            loc.type(value, delay=_random.randint(50, 120))
                            _human_delay(200, 500)
                            break
                    except Exception:
                        continue

            _human_delay(800, 1500)

            # ── 9. Нажимаем «Показать» ───────────────────────────────────────
            log.info("[Browser] Нажимаем «Показать»...")
            for show_sel in [
                "button:has-text('Показать')",
                "input[value='Показать']",
                "text=Показать",
            ]:
                try:
                    loc = page.locator(show_sel).first
                    if loc.is_visible(timeout=3000):
                        _human_delay(600, 1200)
                        _human_move_and_click(page, loc)
                        break
                except Exception:
                    continue

            _human_delay(4000, 7000)   # ждём построения отчёта
            _human_scroll(page)
            _human_delay(1000, 2000)

            # ── 10. Скачиваем XLSX ────────────────────────────────────────────
            log.info("[Browser] Ищем кнопку скачивания XLSX...")
            for dl_sel in [
                "a[href*='.xlsx']",
                "button:has-text('Скачать')",
                "a:has-text('Скачать')",
                "text=xlsx",
                "text=Excel",
                "text=Экспорт",
                "[class*='download']",
            ]:
                try:
                    loc = page.locator(dl_sel).first
                    if not loc.is_visible(timeout=3000):
                        continue
                    _human_delay(600, 1500)
                    _human_move_and_click(page, loc)
                    _human_delay(500, 1000)

                    # Если появилось выпадающее меню — ищем XLSX пункт
                    for xlsx_sel in ["text=xlsx", "text=Excel", "text=.xlsx"]:
                        try:
                            xls_loc = page.locator(xlsx_sel).first
                            if xls_loc.is_visible(timeout=2000):
                                with page.expect_download(timeout=30_000) as dl_info:
                                    _human_move_and_click(page, xls_loc)
                                dl = dl_info.value
                                path = os.path.join(tmpdir, dl.suggested_filename or "report.xlsx")
                                dl.save_as(path)
                                downloaded_file = path
                                log.info("[Browser] Файл скачан: %s", dl.suggested_filename)
                                break
                        except Exception:
                            continue

                    if downloaded_file:
                        break

                    # Прямая загрузка после одного клика
                    try:
                        with page.expect_download(timeout=15_000) as dl_info:
                            pass
                        dl = dl_info.value
                        path = os.path.join(tmpdir, dl.suggested_filename or "report.xlsx")
                        dl.save_as(path)
                        downloaded_file = path
                        log.info("[Browser] Файл скачан (прямая загрузка): %s", dl.suggested_filename)
                        break
                    except Exception:
                        pass

                except Exception:
                    continue

        except Exception as exc:
            log.error("[Browser] Неожиданная ошибка: %s", exc, exc_info=True)
        finally:
            context.close()
            browser.close()

        if not downloaded_file or not os.path.exists(downloaded_file):
            log.error("[Browser] Не удалось скачать отчёт из Яндекс.Директ.")
            return None

        return _parse_direct_xlsx(downloaded_file)


def _parse_direct_xlsx(filepath: str) -> float | None:
    """
    Парсит XLSX-файл статистики Яндекс.Директ.
    Ищет колонку «Стоимость» / «Cost» / «Расход» и суммирует значения.
    Возвращает итоговую сумму расходов в рублях или None при ошибке.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        COST_KEYWORDS = {"стоимость", "cost", "расход", "затраты", "сумма"}
        cost_col_idx  = None
        header_row_idx = None

        for row_idx, row in enumerate(ws.iter_rows(max_row=20), start=1):
            for col_idx, cell in enumerate(row):
                if cell.value and str(cell.value).strip().lower() in COST_KEYWORDS:
                    cost_col_idx  = col_idx
                    header_row_idx = row_idx
                    log.info("[Parser] Найдена колонка расходов '%s' (строка %d, кол. %d)",
                             cell.value, row_idx, col_idx + 1)
                    break
            if cost_col_idx is not None:
                break

        if cost_col_idx is None:
            log.warning("[Parser] Колонка расходов не найдена по ключевым словам.")
            return None

        total = 0.0
        for row in ws.iter_rows(min_row=(header_row_idx or 1) + 1):
            cell = row[cost_col_idx]
            if cell.value is None:
                continue
            try:
                total += float(
                    str(cell.value)
                    .replace(",", ".")
                    .replace("\xa0", "")
                    .replace(" ", "")
                )
            except (TypeError, ValueError):
                continue

        wb.close()
        log.info("[Parser] Суммарные расходы из XLSX: %.2f руб.", total)
        return total

    except Exception as exc:
        log.error("[Parser] Ошибка при разборе XLSX: %s", exc, exc_info=True)
        return None


def build_update_requests(
    sheet_id: int,
    header_row: int,
    windows: dict,
    aggregated: dict,
) -> list[dict]:
    """
    Формирует список batchUpdate requests для записи данных в Google Sheets.
    """
    # ── Маппинг столбцов (0-indexed) ────────────────────────────────────────
    COL_SOURCE   = 0   # Столбец A — "Источник"
    COL_WED      = 1   # Столбец B — Среда
    COL_THU      = 2   # Столбец C — Четверг
    COL_FRI      = 3   # Столбец D — Пятница
    COL_MON      = 4   # Столбец E — Понедельник (Сб+Вс+Пн)
    COL_TUE      = 5   # Столбец F — Вторник
    COL_TOTAL    = 6   # Столбец G — "Итого за неделю"
    COL_TARGETED = 7   # Столбец H — "Кол-во целевых"
    COL_BUDGET   = 8   # Столбец I — "Израсходованный бюджет" (ручной ввод / авто)
    COL_AVG_LEAD = 9   # Столбец J — "Общая цена лида" (формула)
    COL_TGT_LEAD = 10  # Столбец K — "Цена целевого лида" (формула)

    daily = aggregated["daily"]
    targeted = aggregated["targeted"]
    budgets = aggregated.get("budgets", {})

    # Автоматический расчет еженедельной доли бюджета на SEO barssport.com (30 000 руб. в месяц)
    current_wed = windows["current_wed"]
    m_num = current_wed.month
    y_num = current_wed.year
    
    # Считаем количество сред в этом месяце
    num_weeks = 0
    from datetime import date, timedelta
    d = date(y_num, m_num, 1)
    while d.month == m_num:
        if d.weekday() == 2:  # Среда
            num_weeks += 1
        d += timedelta(days=1)
        
    if num_weeks > 0:
        seo_weekly = round(30000.0 / num_weeks, 2)
        if "SEO barssport.com" not in budgets or budgets["SEO barssport.com"] is None:
            budgets["SEO barssport.com"] = seo_weekly

    requests_list = []

    def cell_update(row_0idx: int, col_0idx: int, value) -> dict:
        return {
            "updateCells": {
                "rows": [{
                    "values": [{
                        "userEnteredValue": (
                            {"numberValue": value}
                            if isinstance(value, (int, float))
                            else {"stringValue": str(value)}
                        )
                    }]
                }],
                "fields": "userEnteredValue",
                "start": {
                    "sheetId": sheet_id,
                    "rowIndex": row_0idx,
                    "columnIndex": col_0idx,
                },
            }
        }

    def formula_update(row_0idx: int, col_0idx: int, formula: str) -> dict:
        return {
            "updateCells": {
                "rows": [{
                    "values": [{
                        "userEnteredValue": {"formulaValue": formula}
                    }]
                }],
                "fields": "userEnteredValue",
                "start": {
                    "sheetId": sheet_id,
                    "rowIndex": row_0idx,
                    "columnIndex": col_0idx,
                },
            }
        }

    # Суммируем по каждому источнику
    for row_offset, source_name in enumerate(TABLE_ROWS):
        # header_row - это 1-indexed строка подзаголовков (с датами).
        # Data-строки начинаются сразу под ней (т.е. на индексе header_row в 0-indexed системе).
        data_row_0idx = header_row + row_offset

        requests_list.append(cell_update(data_row_0idx, COL_SOURCE, source_name))

        row_data = daily.get(source_name, {})

        # Подготовка данных по дням
        val_wed = row_data.get("Ср", 0)
        val_thu = row_data.get("Чт", 0)
        val_fri = row_data.get("Пт", 0)
        val_mon = row_data.get("Сб", 0) + row_data.get("Вс", 0) + row_data.get("Пн", 0)
        val_tue = row_data.get("Вт", 0)

        requests_list.append(cell_update(data_row_0idx, COL_WED, val_wed))
        requests_list.append(cell_update(data_row_0idx, COL_THU, val_thu))
        requests_list.append(cell_update(data_row_0idx, COL_FRI, val_fri))
        requests_list.append(cell_update(data_row_0idx, COL_MON, val_mon))
        requests_list.append(cell_update(data_row_0idx, COL_TUE, val_tue))

        # Итого за неделю по лидам (сумма дней)
        total_leads = val_wed + val_thu + val_fri + val_mon + val_tue
        requests_list.append(cell_update(data_row_0idx, COL_TOTAL, total_leads))

        # Кол-во целевых
        tgt = targeted.get(source_name, 0)
        requests_list.append(cell_update(data_row_0idx, COL_TARGETED, tgt))

        # Автоматический бюджет (если есть в переданных бюджетах)
        if source_name in budgets and budgets[source_name] is not None:
            requests_list.append(cell_update(data_row_0idx, COL_BUDGET, budgets[source_name]))

        # Формулы (с разделителем ; для русской локали)
        r = data_row_0idx + 1

        # Общая цена лида (Col J)
        requests_list.append(formula_update(
            data_row_0idx, COL_AVG_LEAD,
            f'=IF(G{r}>0; I{r}/G{r}; "")'
        ))

        # Цена целевого лида (Col K)
        requests_list.append(formula_update(
            data_row_0idx, COL_TGT_LEAD,
            f'=IF(H{r}>0; I{r}/H{r}; "")'
        ))

    # Для строки ИТОГО
    itog_row_0idx = header_row + len(TABLE_ROWS)
    requests_list.append(cell_update(itog_row_0idx, COL_SOURCE, "ИТОГО"))

    # ИТОГО считает через формулы SUM и IF, прописанные на шаге создания шаблона,
    # поэтому нам не нужно вручную перезаписывать ячейки строки ИТОГО.

    log.info("Подготовлено %d requests для Google Sheets", len(requests_list))
    return requests_list


def write_to_google_sheets(
    sheets,
    spreadsheet_id: str,
    tab_name: str,
    sheet_id: int,
    requests_list: list[dict],
):
    """Отправляет данные в Google Sheets через batchUpdate."""
    if not requests_list:
        log.warning("Нет данных для записи в Google Sheets.")
        return

    body = {"requests": requests_list}
    sheets.batchUpdate(spreadsheetId=spreadsheet_id, body=body).execute()
    log.info(
        "✅ Данные успешно записаны в таблицу '%s', вкладка '%s'.",
        spreadsheet_id, tab_name,
    )


def update_dashboard(sheets, spreadsheet_id: str):
    """
    Создает или обновляет аналитический дашборд руководителя в первой вкладке ('Дашборд').
    Анализирует все доступные листы месяцев, суммирует данные и строит сводные таблицы
    с динамическими формулами и профессиональным оформлением.
    """
    log.info("Обновление аналитического дашборда...")

    # 1. Получаем список всех листов
    meta = sheets.get(spreadsheetId=spreadsheet_id).execute()
    all_sheets = meta.get("sheets", [])

    # 2. Фильтруем и сортируем ежемесячные листы (например, "Май 2026")
    MONTHS_RU = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
    MONTH_MAP = {name: idx + 1 for idx, name in enumerate(MONTHS_RU)}

    months_data = []  # Список кортежей: (название_листа, год, номер_месяца, кол_во_недель)
    for s in all_sheets:
        title = s["properties"]["title"]
        parts = title.split()
        if len(parts) == 2 and parts[0] in MONTH_MAP and parts[1].isdigit():
            m_name = parts[0]
            year = int(parts[1])
            month_num = MONTH_MAP[m_name]

            # Вычисляем количество сред в этом месяце (количество недель в шаблоне)
            num_weeks = 0
            d = date(year, month_num, 1)
            while d.month == month_num:
                if d.weekday() == 2:  # Среда
                    num_weeks += 1
                d += timedelta(days=1)

            months_data.append((title, year, month_num, num_weeks))

    # Сортируем в хронологическом порядке
    months_data.sort(key=lambda x: (x[1], x[2]))

    if not months_data:
        log.warning("Не найдено ежемесячных листов для формирования дашборда.")
        return

    log.info("Найдены ежемесячные листы для дашборда: %s", [m[0] for m in months_data])

    # 3. Создаем или очищаем вкладку "Дашборд"
    sheet_id = get_or_create_sheet_tab(sheets, spreadsheet_id, "Дашборд")
    
    # Очищаем содержимое перед записью
    sheets.values().clear(spreadsheetId=spreadsheet_id, range="'Дашборд'!A1:Z100").execute()

    num_months = len(months_data)
    num_channels = len(TABLE_ROWS)

    # Вычисляем общее количество строк на дашборде
    # 0: Заголовок дашборда
    # 1: Отступ
    # 2: Заголовки KPI-карточек
    # 3: Значения KPI-карточек
    # 4: Отступ
    # 5: Заголовок Секции 1 (По месяцам)
    # 6: Шапка Таблицы 1
    # 7 .. 7+num_months-1: Данные по месяцам
    # 7+num_months: ИТОГО по месяцам
    # 8+num_months .. 9+num_months: Отступы
    # 10+num_months: Заголовок Секции 2 (По каналам)
    # 11+num_months: Шапка Таблицы 2
    # 12+num_months .. 12+num_months+num_channels-1: Данные по каналам
    # 12+num_months+num_channels: ИТОГО по каналам
    row_count = 13 + num_months + num_channels
    values = [["" for _ in range(9)] for _ in range(row_count)]
    requests_list = []

    # Сброс всех существующих объединений и стилей
    requests_list.append({
        "unmergeCells": {
            "range": {
                "sheetId": sheet_id
            }
        }
    })
    requests_list.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id
            },
            "cell": {
                "userEnteredFormat": {}
            },
            "fields": "userEnteredFormat"
        }
    })

    # Перемещаем дашборд на первую вкладку (индекс 0)
    requests_list.append({
        "updateSheetProperties": {
            "properties": {
                "sheetId": sheet_id,
                "index": 0
            },
            "fields": "index"
        }
    })

    # ── 4. Формирование структуры данных ─────────────────────────────────────
    # Заголовок дашборда (строка 1 / индекс 0)
    values[0][0] = "АНАЛИТИЧЕСКИЙ ДАШБОРД РУКОВОДИТЕЛЯ"

    # KPI Карточки: Заголовки (строка 3 / индекс 2)
    values[2][1] = "Инвестировано всего"
    values[2][2] = "Всего лидов"
    values[2][3] = "Целевых лидов"
    values[2][4] = "Конверсия в целевой"
    values[2][5] = "Цена целевого лида"

    # KPI Карточки: Значения (строка 4 / индекс 3)
    # Формулы ссылаются на строку ИТОГО первой таблицы (эффективность по месяцам)
    r_total_m_row = 7 + num_months + 1  # 1-indexed номер строки ИТОГО таблицы месяцев
    values[3][1] = f"=B{r_total_m_row}"
    values[3][2] = f"=C{r_total_m_row}"
    values[3][3] = f"=D{r_total_m_row}"
    values[3][4] = f'=IF(C4>0; D4/C4; "")'
    values[3][5] = f'=IF(D4>0; B4/D4; "")'

    # Раздел 1: Эффективность по месяцам
    values[5][0] = "ЭФФЕКТИВНОСТЬ ПО МЕСЯЦАМ"
    values[6] = [
        "Месяц", "Бюджет, ₽", "Всего лидов", "Целевых лидов", 
        "Конверсия в целевой, %", "Цена лида, ₽", "Цена целевого лида, ₽", 
        "Динамика целевых, %", "Динамика стоимости, %"
    ]

    for m_idx, (m_title, m_yr, m_mn, n_w) in enumerate(months_data):
        r = 7 + m_idx
        r_num = r + 1

        # Формулы для суммирования показателей по всем неделям месяца:
        # Строка ИТОГО недели в шаблоне: 19 + idx * 28 (бюджет I, всего лидов G, целевых H)
        budget_cells = [f"'{m_title}'!I{19 + idx * 28}" for idx in range(n_w)]
        leads_cells = [f"'{m_title}'!G{19 + idx * 28}" for idx in range(n_w)]
        target_cells = [f"'{m_title}'!H{19 + idx * 28}" for idx in range(n_w)]

        values[r][0] = m_title
        values[r][1] = f"=SUM({'; '.join(budget_cells)})"
        values[r][2] = f"=SUM({'; '.join(leads_cells)})"
        values[r][3] = f"=SUM({'; '.join(target_cells)})"
        values[r][4] = f'=IF(C{r_num}>0; D{r_num}/C{r_num}; "")'
        values[r][5] = f'=IF(C{r_num}>0; B{r_num}/C{r_num}; "")'
        values[r][6] = f'=IF(D{r_num}>0; B{r_num}/D{r_num}; "")'

        # Вычисление Month-over-Month динамики
        if m_idx == 0:
            values[r][7] = ""
            values[r][8] = ""
        else:
            prev_r_num = r_num - 1
            values[r][7] = f'=IF(D{prev_r_num}>0; (D{r_num}-D{prev_r_num})/D{prev_r_num}; "")'
            values[r][8] = f'=IF(G{prev_r_num}>0; (G{r_num}-G{prev_r_num})/G{prev_r_num}; "")'

    # Строка ИТОГО таблицы по месяцам
    r_itogo_m = 7 + num_months
    r_itogo_m_num = r_itogo_m + 1
    values[r_itogo_m][0] = "ИТОГО"
    values[r_itogo_m][1] = f"=SUM(B8:B{r_itogo_m})"
    values[r_itogo_m][2] = f"=SUM(C8:C{r_itogo_m})"
    values[r_itogo_m][3] = f"=SUM(D8:D{r_itogo_m})"
    values[r_itogo_m][4] = f'=IF(C{r_itogo_m_num}>0; D{r_itogo_m_num}/C{r_itogo_m_num}; "")'
    values[r_itogo_m][5] = f'=IF(C{r_itogo_m_num}>0; B{r_itogo_m_num}/C{r_itogo_m_num}; "")'
    values[r_itogo_m][6] = f'=IF(D{r_itogo_m_num}>0; B{r_itogo_m_num}/D{r_itogo_m_num}; "")'
    values[r_itogo_m][7] = ""
    values[r_itogo_m][8] = ""

    # Раздел 2: Эффективность по каналам за все время
    r_sec2_hdr = r_itogo_m + 3
    values[r_sec2_hdr][0] = "ЭФФЕКТИВНОСТЬ ПО КАНАЛАМ (ЗА ВСЕ ВРЕМЯ)"

    r_sec2_tbl_hdr = r_sec2_hdr + 1
    values[r_sec2_tbl_hdr] = [
        "Источник", "Бюджет, ₽", "Всего лидов", "Целевых лидов", 
        "Конверсия в целевой, %", "Цена лида, ₽", "Цена целевого лида, ₽",
        "Доля бюджета, %", "Доля целевых, %"
    ]

    # Строки данных по каналам
    start_c_row = r_sec2_tbl_hdr + 1
    r_itogo_c = start_c_row + num_channels
    r_itogo_c_num = r_itogo_c + 1

    for s_idx, src in enumerate(TABLE_ROWS):
        r = start_c_row + s_idx
        r_num = r + 1

        # Формулы суммирования ячеек конкретного канала по всем неделям всех месяцев:
        # Строка канала s_idx в неделе шаблона: 9 + idx * 28 + s_idx (бюджет J, всего лидов G, целевых H)
        budget_cells = []
        leads_cells = []
        target_cells = []

        for m_title, m_yr, m_mn, n_w in months_data:
            for idx in range(n_w):
                row_in_sheet = 9 + idx * 28 + s_idx
                budget_cells.append(f"'{m_title}'!I{row_in_sheet}")
                leads_cells.append(f"'{m_title}'!G{row_in_sheet}")
                target_cells.append(f"'{m_title}'!H{row_in_sheet}")

        values[r][0] = src
        values[r][1] = f"=SUM({'; '.join(budget_cells)})"
        values[r][2] = f"=SUM({'; '.join(leads_cells)})"
        values[r][3] = f"=SUM({'; '.join(target_cells)})"
        values[r][4] = f'=IF(C{r_num}>0; D{r_num}/C{r_num}; "")'
        values[r][5] = f'=IF(C{r_num}>0; B{r_num}/C{r_num}; "")'
        values[r][6] = f'=IF(D{r_num}>0; B{r_num}/D{r_num}; "")'

        # Вычисление долей бюджета и целевых лидов по каналам
        values[r][7] = f'=IF(B{r_itogo_c_num}>0; B{r_num}/B{r_itogo_c_num}; "")'
        values[r][8] = f'=IF(D{r_itogo_c_num}>0; D{r_num}/D{r_itogo_c_num}; "")'

    # Строка ИТОГО таблицы по каналам
    values[r_itogo_c][0] = "ИТОГО"
    values[r_itogo_c][1] = f"=SUM(B{start_c_row+1}:B{r_itogo_c})"
    values[r_itogo_c][2] = f"=SUM(C{start_c_row+1}:C{r_itogo_c})"
    values[r_itogo_c][3] = f"=SUM(D{start_c_row+1}:D{r_itogo_c})"
    values[r_itogo_c][4] = f'=IF(C{r_itogo_c_num}>0; D{r_itogo_c_num}/C{r_itogo_c_num}; "")'
    values[r_itogo_c][5] = f'=IF(C{r_itogo_c_num}>0; B{r_itogo_c_num}/C{r_itogo_c_num}; "")'
    values[r_itogo_c][6] = f'=IF(D{r_itogo_c_num}>0; B{r_itogo_c_num}/D{r_itogo_c_num}; "")'
    values[r_itogo_c][7] = ""
    values[r_itogo_c][8] = ""

    # ── 5. Оформление и Стилизирование (batchUpdate) ─────────────────────────
    def add_merge(s_r, e_r, s_c, e_c):
        requests_list.append({
            "mergeCells": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": s_r,
                    "endRowIndex": e_r + 1,
                    "startColumnIndex": s_c,
                    "endColumnIndex": e_c + 1,
                },
                "mergeType": "MERGE_ALL",
            }
        })

    def format_range(s_r, e_r, s_c, e_c, bg_hex=None, fg_hex=None, size=9, bold=False, italic=False, align="CENTER", num_pattern=None):
        cell_format = {}
        fields = []

        if bg_hex:
            cell_format["backgroundColor"] = hex_to_rgb(bg_hex)
            fields.append("backgroundColor")

        text_fmt = {}
        if fg_hex:
            text_fmt["foregroundColor"] = hex_to_rgb(fg_hex)
        text_fmt["fontSize"] = size
        text_fmt["bold"] = bold
        text_fmt["italic"] = italic
        cell_format["textFormat"] = text_fmt
        fields.append("textFormat")

        cell_format["horizontalAlignment"] = align
        cell_format["verticalAlignment"] = "MIDDLE"
        fields.extend(["horizontalAlignment", "verticalAlignment"])

        if num_pattern:
            cell_format["numberFormat"] = {
                "type": "CURRENCY" if "₽" in num_pattern else ("PERCENT" if "%" in num_pattern else "NUMBER"),
                "pattern": num_pattern
            }
            fields.append("numberFormat")

        border = make_border("#CBD5E1")
        cell_format["borders"] = {
            "top": border,
            "bottom": border,
            "left": border,
            "right": border
        }
        fields.append("borders")

        requests_list.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": s_r,
                    "endRowIndex": e_r + 1,
                    "startColumnIndex": s_c,
                    "endColumnIndex": e_c + 1,
                },
                "cell": {
                    "userEnteredFormat": cell_format
                },
                "fields": f"userEnteredFormat({','.join(fields)})"
            }
        })

    # Стилизация заголовка дашборда
    add_merge(0, 0, 0, 8)
    format_range(0, 0, 0, 8, bg_hex="#1E293B", fg_hex="#FFFFFF", size=13, bold=True)

    # Стилизация KPI-карточек
    format_range(2, 2, 1, 5, bg_hex="#475569", fg_hex="#FFFFFF", size=9, bold=True)
    format_range(3, 3, 1, 5, bg_hex="#F8FAFC", size=14, bold=True)

    # Форматы чисел в KPI карточках
    format_range(3, 3, 1, 1, bg_hex="#F8FAFC", size=14, bold=True, num_pattern="#,##0\" ₽\"")
    format_range(3, 3, 2, 3, bg_hex="#F8FAFC", size=14, bold=True, num_pattern="#,##0")
    format_range(3, 3, 4, 4, bg_hex="#F8FAFC", size=14, bold=True, num_pattern="0.0%")
    format_range(3, 3, 5, 5, bg_hex="#F8FAFC", size=14, bold=True, num_pattern="#,##0\" ₽\"")

    # Секция 1: Заголовок и шапка таблицы по месяцам
    add_merge(5, 5, 0, 8)
    format_range(5, 5, 0, 8, bg_hex="#475569", fg_hex="#FFFFFF", size=11, bold=True, align="LEFT")
    format_range(6, 6, 0, 8, bg_hex="#334155", fg_hex="#FFFFFF", size=10, bold=True)

    # Данные таблицы по месяцам
    for m_idx in range(num_months):
        r = 7 + m_idx
        bg = "#FFFFFF" if m_idx % 2 == 0 else "#F8FAFC"
        format_range(r, r, 0, 8, bg_hex=bg, size=9)
        format_range(r, r, 0, 0, bg_hex=bg, size=9, bold=True, align="LEFT")
        format_range(r, r, 1, 1, bg_hex=bg, size=9, align="RIGHT", num_pattern="#,##0\" ₽\"")
        format_range(r, r, 2, 3, bg_hex=bg, size=9, num_pattern="#,##0")
        format_range(r, r, 4, 4, bg_hex=bg, size=9, num_pattern="0.0%")
        format_range(r, r, 5, 6, bg_hex=bg, size=9, align="RIGHT", num_pattern="#,##0\" ₽\"")
        format_range(r, r, 7, 8, bg_hex=bg, size=9, num_pattern="0.0%")

    # ИТОГО таблицы по месяцам
    format_range(r_itogo_m, r_itogo_m, 0, 8, bg_hex="#E2E8F0", size=10, bold=True)
    format_range(r_itogo_m, r_itogo_m, 0, 0, bg_hex="#E2E8F0", size=10, bold=True, align="LEFT")
    format_range(r_itogo_m, r_itogo_m, 1, 1, bg_hex="#E2E8F0", size=10, bold=True, align="RIGHT", num_pattern="#,##0\" ₽\"")
    format_range(r_itogo_m, r_itogo_m, 2, 3, bg_hex="#E2E8F0", size=10, bold=True, num_pattern="#,##0")
    format_range(r_itogo_m, r_itogo_m, 4, 4, bg_hex="#E2E8F0", size=10, bold=True, num_pattern="0.0%")
    format_range(r_itogo_m, r_itogo_m, 5, 6, bg_hex="#E2E8F0", size=10, bold=True, align="RIGHT", num_pattern="#,##0\" ₽\"")
    format_range(r_itogo_m, r_itogo_m, 7, 8, bg_hex="#E2E8F0", size=10, bold=True, num_pattern="0.0%")

    # Секция 2: Заголовок и шапка таблицы по каналам
    add_merge(r_sec2_hdr, r_sec2_hdr, 0, 8)
    format_range(r_sec2_hdr, r_sec2_hdr, 0, 8, bg_hex="#475569", fg_hex="#FFFFFF", size=11, bold=True, align="LEFT")
    format_range(r_sec2_tbl_hdr, r_sec2_tbl_hdr, 0, 8, bg_hex="#334155", fg_hex="#FFFFFF", size=10, bold=True)

    # Данные таблицы по каналам
    for s_idx in range(num_channels):
        r = start_c_row + s_idx
        bg = "#FFFFFF" if s_idx % 2 == 0 else "#F8FAFC"
        format_range(r, r, 0, 8, bg_hex=bg, size=9)
        format_range(r, r, 0, 0, bg_hex=bg, size=9, bold=True, align="LEFT")
        format_range(r, r, 1, 1, bg_hex=bg, size=9, align="RIGHT", num_pattern="#,##0\" ₽\"")
        format_range(r, r, 2, 3, bg_hex=bg, size=9, num_pattern="#,##0")
        format_range(r, r, 4, 4, bg_hex=bg, size=9, num_pattern="0.0%")
        format_range(r, r, 5, 6, bg_hex=bg, size=9, align="RIGHT", num_pattern="#,##0\" ₽\"")
        format_range(r, r, 7, 8, bg_hex=bg, size=9, num_pattern="0.0%")

    # ИТОГО таблицы по каналам
    format_range(r_itogo_c, r_itogo_c, 0, 8, bg_hex="#E2E8F0", size=10, bold=True)
    format_range(r_itogo_c, r_itogo_c, 0, 0, bg_hex="#E2E8F0", size=10, bold=True, align="LEFT")
    format_range(r_itogo_c, r_itogo_c, 1, 1, bg_hex="#E2E8F0", size=10, bold=True, align="RIGHT", num_pattern="#,##0\" ₽\"")
    format_range(r_itogo_c, r_itogo_c, 2, 3, bg_hex="#E2E8F0", size=10, bold=True, num_pattern="#,##0")
    format_range(r_itogo_c, r_itogo_c, 4, 4, bg_hex="#E2E8F0", size=10, bold=True, num_pattern="0.0%")
    format_range(r_itogo_c, r_itogo_c, 5, 6, bg_hex="#E2E8F0", size=10, bold=True, align="RIGHT", num_pattern="#,##0\" ₽\"")
    format_range(r_itogo_c, r_itogo_c, 7, 8, bg_hex="#E2E8F0", size=10, bold=True, num_pattern="0.0%")

    # ── 6. Размеры ячеек (Ширина колонок) ────────────────────────────────────
    col_widths = {
        0: 220,  # Месяц / Источник
        1: 130,  # Бюджет
        2: 110,  # Всего лидов
        3: 110,  # Целевых лидов
        4: 150,  # Конверсия в целевой, %
        5: 140,  # Цена лида, ₽
        6: 150,  # Цена целевого лида, ₽
        7: 150,  # Динамика целевых / Доля бюджета, %
        8: 150,  # Динамика стоимости / Доля целевых, %
    }
    for col_idx, width in col_widths.items():
        requests_list.append({
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": col_idx,
                    "endIndex": col_idx + 1
                },
                "properties": {"pixelSize": width},
                "fields": "pixelSize"
            }
        })

    # Высота строк
    for r in range(row_count):
        h = 25
        if r == 0:
            h = 50
        elif r == 2:
            h = 24
        elif r == 3:
            h = 35
        elif r in (5, 6, r_sec2_hdr, r_sec2_tbl_hdr):
            h = 30
        elif r in (1, 4, r_itogo_m + 1, r_itogo_m + 2):
            h = 15

        requests_list.append({
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": r,
                    "endIndex": r + 1
                },
                "properties": {"pixelSize": h},
                "fields": "pixelSize"
            }
        })

    # ── 7. Выполнение запросов на запись ─────────────────────────────────────
    # Записываем формулы и тексты в ячейки
    range_name = f"'Дашборд'!A1"
    body = {"values": values}
    sheets.values().update(
        spreadsheetId=spreadsheet_id,
        range=range_name,
        valueInputOption="USER_ENTERED",
        body=body,
    ).execute()

    # Запускаем batchUpdate форматирования
    if requests_list:
        sheets.batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": requests_list},
        ).execute()

    log.info("✅ Аналитический дашборд успешно обновлен!")


# ---------------------------------------------------------------------------
# Главная функция
# ---------------------------------------------------------------------------

def process_week(sheets, webhook_url: str, spreadsheet_id: str, target_wed: date):
    """Выполняет полный цикл обработки и обновления данных за одну отчетную неделю."""
    log.info("=" * 60)
    log.info("ОБРАБОТКА НЕДЕЛИ С: %s", target_wed.strftime("%d.%m.%Y"))
    log.info("=" * 60)

    # 1. Вычисляем временные окна для конкретной среды
    windows = compute_date_windows(target_wed)
    current_wed = windows["current_wed"]
    current_tue = windows["current_tue"]

    # 2. Запрашиваем лиды из CRM за этот период (только Ср-Вт этой недели)
    all_leads = fetch_leads(webhook_url, current_wed, current_tue)

    # 3. Агрегируем данные (целевые считаются строго по дате создания)
    aggregated = aggregate_leads(windows, all_leads)

    # 4. Получаем расходы Яндекс.Директ
    aggregated["budgets"] = {}
    yandex_token    = os.environ.get("YANDEX_DIRECT_TOKEN", "").strip()
    yandex_login    = os.environ.get("YANDEX_DIRECT_CLIENT_LOGIN", "").strip()
    yandex_browser_login    = os.environ.get("YANDEX_LOGIN", "").strip()
    yandex_browser_password = os.environ.get("YANDEX_PASSWORD", "").strip()
    yandex_secret_answer    = os.environ.get("YANDEX_SECRET_ANSWER", "").strip() or None
    
    yandex_cost = None

    # Попытка 0: CSV-файл(ы) из репозитория
    csv_cost = parse_direct_csv_files(current_wed, current_tue)
    if csv_cost is not None:
        yandex_cost = csv_cost

    # Попытка 1: официальный API
    if yandex_cost is None and yandex_token:
        try:
            log.info("Пробуем получить расходы через API Яндекс.Директ...")
            yandex_cost = fetch_yandex_direct_cost(yandex_token, yandex_login or None, current_wed, current_tue)
        except Exception as e:
            log.warning("API Яндекс.Директ недоступен: %s. Переключаемся на браузер.", e)
            yandex_cost = None

    # Попытка 2: браузерная автоматизация
    if yandex_cost is None and yandex_browser_login and yandex_browser_password:
        try:
            log.info("Получаем расходы через браузерную автоматизацию...")
            yandex_cost = fetch_yandex_direct_cost_browser(
                yandex_browser_login,
                yandex_browser_password,
                yandex_secret_answer,
                current_wed,
                current_tue,
            )
        except Exception as e:
            log.error("Ошибка браузерного импорта расходов Яндекс.Директ: %s", e, exc_info=True)

    if yandex_cost is not None:
        aggregated["budgets"]["Я.Директ e-17479930"] = yandex_cost
        log.info("Расходы Яндекс.Директ записаны: %.2f руб.", yandex_cost)
    else:
        log.warning("Расходы Яндекс.Директ не получены ни из CSV, ни через API, ни через браузер.")

    # 5. Работаем с Google Sheets
    MONTH_NAMES_RU = {
        1: "Январь", 2: "Февраль", 3: "Март",
        4: "Апрель", 5: "Май",     6: "Июнь",
        7: "Июль",   8: "Август",  9: "Сентябрь",
        10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
    }
    tab_name = f"{MONTH_NAMES_RU[current_wed.month]} {current_wed.year}"
    log.info("Целевая вкладка: '%s'", tab_name)

    # Получаем или создаём вкладку
    sheet_id = get_or_create_sheet_tab(sheets, spreadsheet_id, tab_name)

    # Ищем блок текущей недели по дате Среды
    header_row = find_week_block_row(sheets, spreadsheet_id, tab_name, current_wed)

    if header_row is None:
        create_month_template(sheets, spreadsheet_id, tab_name, sheet_id, current_wed)
        header_row = find_week_block_row(sheets, spreadsheet_id, tab_name, current_wed)

    if header_row is None:
        log.error(
            "❌ Не удалось найти или создать блок недели для даты %s на листе '%s'.",
            current_wed, tab_name
        )
        return

    # 6. Формируем и отправляем запросы к Sheets API
    requests_list = build_update_requests(sheet_id, header_row, windows, aggregated)
    write_to_google_sheets(sheets, spreadsheet_id, tab_name, sheet_id, requests_list)


def main():
    log.info("=" * 60)
    log.info("Запуск скрипта еженедельного отчёта по лидам")
    log.info("=" * 60)

    # 1. Читаем переменные среды
    webhook_url          = get_env("BITRIX24_WEBHOOK_URL")
    service_account_json = get_env("GOOGLE_SERVICE_ACCOUNT_JSON")
    spreadsheet_id       = get_env("SPREADSHEET_ID")

    sheets = get_sheets_service(service_account_json)

    # 2. Вычисляем текущую среду отчетной недели
    windows = compute_date_windows()
    current_wed = windows["current_wed"]

    # 3. Обновляем две недели: предыдущую (чтобы окончательно зафиксировать) и текущую
    prev_wed = current_wed - timedelta(days=7)

    log.info("--- ШАГ 1: Обновление ПРЕДЫДУЩЕЙ недели (%s) ---", prev_wed)
    try:
        process_week(sheets, webhook_url, spreadsheet_id, prev_wed)
    except Exception as e:
        log.error("Ошибка при обновлении предыдущей недели %s: %s", prev_wed, e, exc_info=True)

    log.info("--- ШАГ 2: Обновление ТЕКУЩЕЙ недели (%s) ---", current_wed)
    try:
        process_week(sheets, webhook_url, spreadsheet_id, current_wed)
    except Exception as e:
        log.error("Ошибка при обновлении текущей недели %s: %s", current_wed, e, exc_info=True)

    # 4. Обновляем аналитический дашборд на первой вкладке
    try:
        update_dashboard(sheets, spreadsheet_id)
    except Exception as e:
        log.error("Ошибка при обновлении дашборда: %s", e, exc_info=True)

    log.info("=" * 60)
    log.info("Скрипт завершён успешно.")
    log.info("=" * 60)


if __name__ == "__main__":
    main()

